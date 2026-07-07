import openpyxl
import json

file_path = r"C:\Users\user\Downloads\복사본 주니어 능률 VOCA_실력_어휘 리스트.xlsx"

print("[OK] 주니어 능률 어휘 변환 시작\n")

# 엑셀 파일 읽기
wb = openpyxl.load_workbook(file_path)
ws = wb.active

words_data = []

# 헤더 스킵 (1행)
for row_idx in range(2, ws.max_row + 1):
    unit = ws.cell(row_idx, 1).value  # DAY 01, 02, ...
    eng_word = ws.cell(row_idx, 2).value  # 영단어
    meanings = ws.cell(row_idx, 3).value  # 뜻

    if not eng_word or not meanings:
        continue

    # 정리
    unit = str(unit).strip() if unit else "unknown"
    eng_word = str(eng_word).strip()
    meanings = str(meanings).strip()

    # 뜻 파싱 (쉼표로 구분)
    meanings_ko = [m.strip() for m in meanings.split(",") if m.strip()]

    if eng_word and meanings_ko:
        words_data.append({
            "text": eng_word,
            "meanings_ko": meanings_ko,
            "unit": unit
        })

print(f"[OK] 추출 완료: {len(words_data)}개 단어\n")

# 샘플 출력
print("=== 샘플 데이터 (첫 20개) ===")
for i, word in enumerate(words_data[:20]):
    print(f"{i+1}. {word['text']}")
    print(f"   {', '.join(word['meanings_ko'][:3])}")

# JSON 파일 저장
output_path = r"C:\Users\user\Downloads\junior-vocab-final.json"
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(words_data, f, ensure_ascii=False, indent=2)

print(f"\n[OK] JSON 저장 완료: {output_path}")
print(f"[OK] 총 {len(words_data)}개 단어")
