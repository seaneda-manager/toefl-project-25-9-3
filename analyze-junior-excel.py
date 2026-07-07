import openpyxl
import json

file_path = r"C:\Users\user\Downloads\복사본 주니어 능률 VOCA_실력_어휘 리스트.xlsx"

# 엑셀 파일 읽기
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"시트명: {ws.title}")
print(f"행: {ws.max_row}, 열: {ws.max_column}\n")

# 헤더 확인
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    headers.append(str(cell_value) if cell_value else f"Col{col}")

print(f"컬럼: {headers}\n")

# 처음 15행 미리보기
print("=== 데이터 샘플 (첫 15행) ===")
for row_idx in range(1, min(16, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx).value
        row_data.append(str(cell) if cell else "")
    print(f"행 {row_idx}: {' | '.join(row_data)}")
